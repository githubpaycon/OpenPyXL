"""
https://openpyxl.readthedocs.io/en/stable/
pip install openpyxl
pipenv install openpyxl
"""

from random import uniform
import openpyxl
import os

# create plan
planilha = openpyxl.Workbook()

# create sheet in plan
#                    plan_name, index (organização)
planilha.create_sheet('Planilha1', 0)
planilha.create_sheet('Planilha2', 1)
# se for index 0 na primeira e 0 na segunda a primeira é empurrada para 1

# get plans
planilha1 = planilha['Planilha1']
planilha2 = planilha['Planilha2']

# adding datas in planilha1
for linha in range(5, 16):
    numero_pedido = linha - 1
    id_produto = 1200 + linha
    preco = round(uniform(10, 100), 2) # cria preços diferentes arredondado de 2 casas decimais

    planilha1.cell(linha, 1).value = numero_pedido
    planilha1.cell(linha, 2).value = id_produto
    planilha1.cell(linha, 3).value = f'R$ {preco}'

# adding datas in planilha2
for linha in range(5, 16):
    numero_pedido = linha - 1
    id_produto = 1200 + linha
    planilha2.cell(linha, column=1).value = f'Gabriel {linha} {round(uniform(10, 100), 2)}'
    planilha2.cell(linha, column=2).value = f'Lopes {linha} {round(uniform(10, 100), 2)}'
    planilha2.cell(linha, column=3).value = f'Souza {linha} {round(uniform(10, 100), 2)}'
    
    
# save archive
planilha.save('nova_plan.xlsx')