"""
https://openpyxl.readthedocs.io/en/stable/
pip install openpyxl
pipenv install openpyxl
"""

import openpyxl
import os
from random import uniform # ger aum numero de ponto flutuante

# nome da planilha (abre a planilha)
pedidos = openpyxl.load_workbook(os.path.abspath('pedidos.xlsx'))
nome_planilhas = pedidos.sheetnames # saber quais são as planilhas que tem no arquivo excel

# print(nome_planilhas) # ['Página1']

# pegar a sheet da planilha pedidos
plan1 = pedidos['Página1']

# alterando dados de uma planilha
# plan1['b3'].value = 2200 ## nao mexe na planilha original

# adicionando dados em colunas = a5, b5, c5, 
#            inicio, Fim linha
for linha in range(5, 16):
    numero_pedido = linha - 1
    id_produto = 1200 + linha
    preco = round(uniform(10, 100), 2) # cria preços diferentes arredondado de 2 casas decimais

    plan1.cell(linha, 1).value = numero_pedido
    plan1.cell(linha, 2).value = id_produto
    plan1.cell(linha, 3).value = f'R$ {preco}'
    



pedidos.save('new.xlsx')